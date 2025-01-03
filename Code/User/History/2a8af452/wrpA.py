from openpyxl import load_workbook

def extract_every_nth_row(excel_file, csv_file, n=1000):
  """
  Estrae una riga ogni n righe da un file Excel e le salva in un file CSV.

  Args:
    excel_file: Il percorso del file Excel.
    csv_file: Il percorso del file CSV di output.
    n: L'intervallo tra le righe da estrarre.
  """

  # Carica il foglio di lavoro Excel
  workbook = load_workbook(excel_file)
  worksheet = workbook.active

  # Apre il file CSV in modalità scrittura
  with open(csv_file, 'w', newline='') as csvfile:
      # Scrive l'intestazione del CSV (se presente nel foglio Excel)
      csvfile.write(','.join(worksheet[1]))
      csvfile.write('\n')

      # Scrive le righe selezionate nel CSV
      for row in worksheet.iter_rows(min_row=2, values_only=True):
          if not (worksheet.row_dimensions[row[0].row].hidden):  # Salta righe nascoste
              if row[0].row % n == 0:
                  csvfile.write(','.join(str(cell) for cell in row))
                  csvfile.write('\n')

# Esegui lo script
excel_file = 'train_inputs.xlsx'
csv_file = 'train_inputs_filtered.csv'
extract_every_nth_row(excel_file, csv_file)