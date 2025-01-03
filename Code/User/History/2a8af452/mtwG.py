from openpyxl import load_workbook
import os
import sys 


def extract_every_nth_row(excel_file, csv_file, n=1000):
  """
  Estrae una riga ogni n righe da un file Excel e le salva in un file CSV.

  Args:
    excel_file: Il percorso del file Excel.
    csv_file: Il percorso del file CSV di output.
    n: L'intervallo tra le righe da estrarre.
  """

  # Carica il foglio di lavoro Excel
  workbook = load_workbook(excel_file)
  worksheet = workbook.active

  # Apre il file CSV in modalità scrittura
  with open(csv_file, 'w', newline='') as csvfile:
      # Scrive l'intestazione del CSV (se presente nel foglio Excel)
      #csvfile.write(','.join(worksheet[1]))
      #csvfile.write('\n')

      # Scrive le righe selezionate nel CSV
      for row in worksheet.iter_rows(min_row=2, values_only=True):
              #if not (worksheet.row_dimensions[row[0].row].hidden):  # Salta righe nascoste
              if row[0] % n == 0:
                  print("... progress ", row[0])
                  csvfile.write(','.join(str(cell) for cell in row))
                  csvfile.write('\n')

# Esegui lo script
for fname in os.listdir():
    if ".xlsx" in fname:
        print("Filename ", fname)
        excel_file = fname
        csv_file = fname.replace(".xlsx","_filter.csv")
        #csv_file = open(csv_filename,"w")
        extract_every_nth_row(excel_file, csv_file)